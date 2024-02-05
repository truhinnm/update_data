import io
import os
import msoffcrypto
from users_passwords import *
from win32com.client.gencache import EnsureDispatch
from openpyxl import load_workbook
import time


def decrypt_file(file_path, file_password):
    decrypted_workbook = io.BytesIO()
    with open(file_path, 'rb') as cur_file:
        office_file = msoffcrypto.OfficeFile(cur_file)
        office_file.load_key(password=file_password)
        office_file.decrypt(decrypted_workbook)
    return decrypted_workbook


def find_cell(sheet, string_column, target_column, search_string):
    for cell in sheet[string_column]:
        if cell.value is not None:
            if cell.value == search_string:
                return target_column + str(cell.row)


def find_main_row(search_string, sheet, column):
    for cell in sheet[column]:
        if cell.value is not None:
            if " ".join(cell.value.split()[-2:]) == search_string:
                return str(cell.row)
            if cell.value == search_string:
                return str(cell.row)


def edit_workbook(workbook, sheet, file_path, cell, data):
    sheet[cell] = data
    workbook.save(file_path)


def set_wb_pass(file_dir_path, read_password, write_password):
    xl_file = EnsureDispatch("Excel.Application")
    wb = xl_file.Workbooks.Open(file_dir_path)
    xl_file.DisplayAlerts = False
    wb.SaveAs(file_dir_path, Password=read_password, WriteResPassword=write_password)
    wb.Close()
    xl_file.Quit()


def open_main_tables(mt_path):
    tables = {}
    for m_root, m_dirs, m_files in os.walk(mt_path):
        for m_file in m_files:
            if m_file.endswith('.xlsx') and not m_file.startswith('~$'):
                path = os.path.join(m_root, m_file)
                tables[m_file] = decrypt_file(path, main_tables[m_file][0])
    return tables


def transfer_data(parent_wb, parent_sheet, child_wb, filename, target_string):
    formatted_filename = filename[:-5]
    parent_ws = parent_wb[parent_sheet]
    child_ws = child_wb.active
    if target_string == "Исполнительская дисциплина":
        parent_target = find_cell(parent_ws, "E", "D", formatted_filename)
        child_target = find_cell(child_ws, "A", "D", target_string)
        child_ws[child_target] = parent_ws[parent_target].value
    if target_string == "Трудовая дисциплина":
        parent_target = find_cell(parent_ws, "A", "B", formatted_filename)
        child_target = find_cell(child_ws, "A", "D", target_string)
        child_ws[child_target] = parent_ws[parent_target].value
    return None


def transfer_to_main(filename, ot_workbook, ind_workbook, ind_val_workbook):
    formatted_filename = filename[:-5]
    ot_ws = ot_workbook.active
    ind_ws = ind_workbook.active
    ind_val_ws = ind_val_workbook.active
    row = find_main_row(formatted_filename, ot_ws, "B")
    id_target = find_cell(ind_ws, "A", "D", "Исполнительская дисциплина")
    td_target = find_cell(ind_ws, "A", "D", "Трудовая дисциплина")
    or_target = find_cell(ind_val_ws, "A", "D", "Итоговая оценка руководителя")
    print("this is " + str(ind_val_ws[or_target].value))
    ot_ws[str("H" + row)] = ind_ws[td_target].value
    ot_ws[str("I" + row)] = ind_val_ws[or_target].value
    ot_ws[str("J" + row)] = ind_ws[id_target].value


if __name__ == '__main__':

    for i in users:
        print("Добрый день! Высылаю пароль от вашего индивидуального листа KPI: " + users[i][0] + ". Название файла: " + i)
    # # Открываем общие таблицы
    #
    # main_path = r"\\192.168.88.16\kpi\Общая таблица"
    # m_tables = open_main_tables(main_path)
    # id_wb = load_workbook(m_tables["Исполнительская дисциплина.xlsx"], data_only=True)
    # td_wb = load_workbook(m_tables["Трудовая дисциплина.xlsx"], data_only=True)
    # ot_wb = load_workbook(m_tables["KPI общая таблица.xlsx"])
    #
    # # Начинаем обработку индивидуальных листов
    #
    # for root, dirs, files in os.walk(r"\\192.168.88.16\kpi"):
    #     for file in files:
    #         if file.endswith('.xlsx') and not file.startswith('~$'):
    #             if ("KPI_Архив" not in root) & ("ДЕМО" not in root) & ("Общая таблица" not in root):
    #                 filepath = os.path.join(root, file)
    #                 print("working on" + filepath)
    #
    #                 # Открываем индивидуальную таблицу
    #
    #                 cur_wb = decrypt_file(filepath, users[file][0])
    #                 ind_wb = load_workbook(cur_wb)
    #                 ind_val_wb = load_workbook(cur_wb, data_only=True)
    #
    #                 # Переносим данные
    #
    #                 transfer_data(id_wb, "Оценка", ind_wb, file, "Исполнительская дисциплина")
    #                 transfer_data(td_wb, "Статистика", ind_wb, file, "Трудовая дисциплина")
    #                 if file == "Якупов Альберт.xlsx":
    #                     transfer_to_main("Производитель работ Якупов А.А..xlsx", ot_wb, ind_wb, ind_val_wb)
    #                 if file == "Якупов А.А..xlsx":
    #                     transfer_to_main("Инженер ПТО Якупов А.А..xlsx", ot_wb, ind_wb, ind_val_wb)
    #                 transfer_to_main(file, ot_wb, ind_wb, ind_val_wb)
    #
    #                 # Сохраняем индивидуальный лист
    #
    #                 ind_wb.save(filepath)
    #                 set_wb_pass(filepath, users[file][0], users[file][1])
    #                 print("DONE")
    #
    # # Сохраняем общую таблицу
    #
    # ot_wb.save(os.path.join(main_path, "KPI общая таблица.xlsx"))
    # set_wb_pass(os.path.join(main_path, "KPI общая таблица.xlsx"),
    #             main_tables["KPI общая таблица.xlsx"][0],
    #             main_tables["KPI общая таблица.xlsx"][1])
    # print("FINALLY")



